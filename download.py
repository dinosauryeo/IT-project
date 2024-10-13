import csv
from pymongo import MongoClient
from datetime import datetime
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side
import os
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders


TIME = 60
COLUMN_C = 87.67
COLUMN_B = 29.67
COLUMN_A = 11.5
COLUMN_D = 25.67
COLUMN_E = 14.33
COLUMN_G = 24.5
ROW = 9
TITLE = 3
HEADER = 5
SIZE = 11
ROW_HEIGHT = 30
BLUE = "0000FF"
BLACK = "000000"
RED = "FF0000"
GREY = "C0C0C0"
LIGHT_BLUE = "E0FFFF"
HORIZONTAL = 'center'
LEFT = 'left'
TAHOMA = 'Tahoma'


'''change the time format'''
def format_time(time_str):
    return time_str.lstrip('0') if time_str.startswith('0') else time_str


'''calculate class time interval'''
def calculate_duration(start_time, end_time):
    """calculate time interval for lecture/lab/tutorial"""
    time_format = "%H:%M" 
    start = datetime.strptime(start_time, time_format)
    end = datetime.strptime(end_time, time_format)

    # calculate duriation
    duration = end - start
    total_minutes = duration.total_seconds() // TIME
    hours = total_minutes // TIME
    minutes = total_minutes % TIME
    if minutes == 0:
        return (str(hours)+'h').replace('.0', '')
    else:
        return (str(hours)+'h'+ str(minutes) +'min').replace('.0', '')



'''translate csv timetable into excel
with well structured format and meet target output design '''
def csv_to_excel(csv_file, excel_file):
    # create border feature
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    tahoma_font = Font(name = TAHOMA, size = SIZE)
    # create a Excelwritter 
    with pd.ExcelWriter(excel_file, engine = 'openpyxl') as writer:
        workbook = writer.book
        #create new work sheet
        worksheet = workbook.create_sheet(title = 'Timetable') 
        worksheet.column_dimensions['A'].width = COLUMN_A
        worksheet.column_dimensions['B'].width = COLUMN_B
        worksheet.column_dimensions['C'].width = COLUMN_C
        worksheet.column_dimensions['D'].width = COLUMN_D
        worksheet.column_dimensions['E'].width = COLUMN_E
        worksheet.column_dimensions['F'].width = COLUMN_E
        # image path find
        image_directory = os.path.join(os.getcwd(), "templates")
        image_dir = os.path.join(image_directory, "static")
        image_d = os.path.join(image_dir, "images")
        image_path = os.path.join(image_d, 'uniphoto.png')
        # create image
        if os.path.exists(image_path):
            img = Image(image_path)  
            #move image into EXCEL
            worksheet.add_image(img, 'A1') 
        else:
            print(f"Image not found at path: {image_path}")
            return False
        
        # import title
        worksheet.append(['']*1 + ["Victorian Institute of Technology Pty Ltd"])
        worksheet.append(['']*1 + ["ABN: 41 085 128 525 RTO No: 20829 TEQSA ID: PRV14007 CRICOS Provider Code: 02044E"])
        worksheet.append(['']*1 + ["Master of Information Technology and Systems"])
        worksheet.append(['Timetable: ']*1 + ["Venue: 123 & 235 Queens Street, Melbourne"])
        worksheet.append(["Note: (a) This is a Master Timetable. You are required to refer to your Unit Allocation (i.e., your enrolled units) to know which units/sessions are applicable to you."])
        worksheet.append(["(b) Time Table may change in the event of some exigencies.(c) Units have additional consulting sessions (based on unit requirements) which is not reflected below including online guided learning"])
        worksheet.append(["(d) L- Lecture (2hr), P- Practical (1hr), T - Tutorial/Lab/Guided Learning (1hr) (e) TA -Teaching Assistant as Rostered."])
        worksheet.append([])  
        data = pd.read_csv(csv_file)
        worksheet.append(list(data.columns))
        #read data into excel
        for row in data.itertuples(index=False):
            # append each line into excel
            worksheet.append(row) 
        worksheet.merge_cells('B1:F1') 
        worksheet.merge_cells('B2:F2') 
        worksheet.merge_cells('B3:F3') 
        worksheet.merge_cells('B4:F4')
        worksheet.merge_cells('A5:F5')
        worksheet.merge_cells('A6:F6')
        worksheet.merge_cells('A7:F7')
        for row in range(1, ROW - 1):
            if row < TITLE:
                merged_cell = worksheet[f'B{row}']
                merged_cell.alignment = Alignment(horizontal=HORIZONTAL, vertical=HORIZONTAL)
                #blue color
                merged_cell.font = Font(color = BLUE, bold = True, name=TAHOMA,size = SIZE) 
            elif TITLE <= row < HEADER:
                merged_cell = worksheet[f'B{row}']
                merged_cell.alignment = Alignment(horizontal= HORIZONTAL, vertical= HORIZONTAL)
                merged_cell.font = Font(color = BLACK, bold = True, name=TAHOMA,size = SIZE) 
            else:
                #red color
                merged_cell = worksheet[f'A{row}']
                merged_cell.alignment = Alignment(horizontal = LEFT, vertical = HORIZONTAL)
                merged_cell.font = Font(color = RED, bold = True, name=TAHOMA, size = SIZE)    

            
        # grey back ground and font
        header_fill = PatternFill(start_color= GREY , end_color = GREY, fill_type="solid")  
        header_font = Font(bold=True, color = BLACK, name = TAHOMA)
        columns = ['A', 'B', 'C', 'D', 'E', 'F']
        for column in columns:
            cell = worksheet[f'{column}9']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal = HORIZONTAL, vertical = HORIZONTAL)
            cell.border = thin_border
        #set data background into light blue color
        for row_index, row in data.iterrows():
            for col_index in range(len(row)):
                cell = worksheet.cell(row = row_index + ROW + 1, column = col_index + 1)
                cell.fill = PatternFill(start_color = LIGHT_BLUE, end_color = LIGHT_BLUE, fill_type = "solid")
                cell.font = tahoma_font
                cell.border = thin_border
        worksheet.row_dimensions[ROW].height = ROW_HEIGHT
        worksheet['D9'].alignment = Alignment(horizontal = HORIZONTAL, vertical = HORIZONTAL, wrap_text = True)






'''
The method is to download each student timetable file from MongoDB 
transfer into csv with student id _timetable
'''
def download_all(year,semester,campus,folder_prefix,degree_name):
    #setup information required to send the email
    server = 'smtp.gmail.com'
    port = 587
    username = "dinosauryeo@gmail.com"
    password = "jucvnvbkwtgcehjo"
    
    try:
        with smtplib.SMTP(server, port) as server:
            #create connection
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls() 
                
            #login and send the mail
            server.login(username, password)
            
            print("server prepared\n")
            
            client = MongoClient('mongodb+srv://dinosauryeo:6OHYa6vF6YUCk48K@cluster0.dajn796.mongodb.net/')
            db_name = f'{year}_{semester}'
            db = client[db_name]
            
            #get the target collection that stoes the timetable 
            folder_pattern = re.compile(f"^{re.escape(folder_prefix)}.*{re.escape(degree_name)}.*{re.escape(campus)}.*")
            collections = db.list_collection_names()
            collection_name = next((coll for coll in collections if folder_pattern.match(coll)), None)
            
            if collection_name is None:
                print(f"No collection found for the given criteria.")
                return False
        
            collection = db[collection_name]
            
            #preparing information for fetching the gmail address of the student to send the email to
            folder_pattern = re.compile(f"^Students-Enrollment-Details.*{re.escape(degree_name)}.*{re.escape(campus)}.*")
            collections = db.list_collection_names()
            collection_name = next((coll for coll in collections if folder_pattern.match(coll)), None)
            
            if collection_name is None:
                return False;
            
            student_collection = db[collection_name]
            
            days_order = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
            # create timetable file in local path
            download_dir = 'student_timetable'
            os.makedirs(download_dir, exist_ok=True)
            # extract each student timetable
            for document in collection.find():
                student_id = document['StudentID']
                print(f"{student_id}\n")
                
                timetables = document.get('Timetable', [])
                sorted_timetables = sorted(timetables, key=lambda x: (days_order.index(x['Day'].lower()), datetime.strptime(x['From'], '%H:%M')))
                # create each students' csv
                filename = f'{student_id}_timetable.csv'
                file_path = os.path.join(download_dir, filename) 
                with open(file_path, 'w', newline = '', encoding = 'utf-8') as file:
                    fieldnames = ['Day', 'Time', 'Unit', 'Classroom\nLevel/ Room/ Venue', 'Lecturer', 'Delivery Mode']
                    writer = csv.DictWriter(file, fieldnames = fieldnames)
                    writer.writeheader()
                    if sorted_timetables:
                        for timetable in sorted_timetables:
                            row = {
                                'Day': timetable.get('Day', '').capitalize(),
                                'Time': format_time(timetable.get('From', '')) + ' to ' + format_time(timetable.get('To', '')) + "(L + T)",
                                'Unit': timetable.get('SubjectCode', '') + '-' + timetable.get('SubjectName', '') + '(' + calculate_duration(format_time(timetable.get('From', '')) , format_time(timetable.get('To', ''))) +' ' +timetable.get('Title', '') + ')',
                                'Classroom\nLevel/ Room/ Venue': timetable.get('Location', ''),
                                'Lecturer': timetable.get('Name', ''),
                                'Delivery Mode': timetable.get('Mode', '')
                            }
                            writer.writerow(row)
                            #print(f"writing into csv\n")
                    else:
                        print(f"No timetable found for StudentID: {student_id}")
                excel_path = os.path.join(download_dir, f'{student_id}_timetable.xlsx')         
                csv_to_excel(file_path, excel_path)
                print("excel verstion timetable ready\n")
            
                #fetching the student's email
                student_data = student_collection.find_one({"StudentID": int(student_id)})
                student_email = student_data["University Email"]
                
                print("student_data foud\n")
                
                #construct the email body
                msg = MIMEMultipart()
                msg['From'] = username
                msg['To'] = student_email
                msg['Subject'] = "Student's timetable"
                body = "Hi, below is your timetable for the following semester"
                msg.attach(MIMEText(body, 'plain'))

                print(excel_path)
                #attch the excel file
                with open(excel_path, "rb") as attachment:
                    print("starts to encode file\n")
                    # Create a MIMEBase object and set its payload to the file content
                    mime_base = MIMEBase('application', 'octet-stream')
                    mime_base.set_payload(attachment.read())
                        
                    # Encode the payload in base64
                    encoders.encode_base64(mime_base)
                            
                    # Add a header to the attachment
                    mime_base.add_header('Content-Disposition', f'attachment; filename="{excel_path.split("/")[-1]}"')
                            
                    # Attach the Excel file to the message
                    msg.attach(mime_base)
                    print("finish encoding file\n")
                
                print("trying to send email\n")
                server.sendmail(username, student_email, msg.as_string())
                print(f"sent to {student_id}")
                
                #delete the file after it had been sent
                
                os.remove(excel_path)
                os.remove(file_path)
        
    except Exception as e:
        return False
    
    return True
    

def download_one(year, semester, campus, folder_prefix, degree_name, student_id):
    client = MongoClient('mongodb+srv://dinosauryeo:6OHYa6vF6YUCk48K@cluster0.dajn796.mongodb.net/')
    db_name = f'{year}_{semester}'
    db = client[db_name]
    
    # 使用正则表达式寻找符合的集合
    folder_pattern = re.compile(f"^{re.escape(folder_prefix)}.*{re.escape(degree_name)}.*{re.escape(campus)}.*")
    collections = db.list_collection_names()
    collection_name = next((coll for coll in collections if folder_pattern.match(coll)), None)

    if collection_name is None:
        print(f"No collection found for the given criteria.")
        return False
    
    collection = db[collection_name]
    days_order = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']

    # 查找指定 student_id 的时间表
    student_data = collection.find_one({"StudentID": int(student_id)})

    if student_data is None:
        print(f"No timetable found for StudentID: {student_id}")
        return False

    timetables = student_data.get('Timetable', [])
    
    if not timetables:
        print(f"No timetable data available for StudentID: {student_id}")
        return False

    # 排序时间表
    sorted_timetables = sorted(timetables, key=lambda x: (days_order.index(x['Day'].lower()), datetime.strptime(x['From'], '%H:%M')))

    # 创建目录
    download_dir = 'student_timetable'
    os.makedirs(download_dir, exist_ok=True)

    # 创建每个学生的csv文件
    filename = f'{student_id}_timetable.csv'
    file_path = os.path.join(download_dir, filename)

    with open(file_path, 'w', newline='', encoding='utf-8') as file:
        fieldnames = ['Day', 'Time', 'Unit', 'Classroom\nLevel/ Room/ Venue', 'Lecturer', 'Delivery Mode']
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()

        for timetable in sorted_timetables:
            row = {
                'Day': timetable.get('Day', '').capitalize(),
                'Time': format_time(timetable.get('From', '')) + ' to ' + format_time(timetable.get('To', '')) + " (L + T)",
                'Unit': timetable.get('SubjectCode', '') + '-' + timetable.get('SubjectName', '') + ' (' + calculate_duration(format_time(timetable.get('From', '')), format_time(timetable.get('To', ''))) + ' ' + timetable.get('Title', '') + ')',
                'Classroom\nLevel/ Room/ Venue': timetable.get('Location', ''),
                'Lecturer': timetable.get('Name', ''),
                'Delivery Mode': timetable.get('Mode', '')
            }
            writer.writerow(row)

    # 将CSV转换为Excel
    excel_path = os.path.join(download_dir, f'{student_id}_timetable.xlsx')
    csv_to_excel(file_path, excel_path)
    
    #fetching the gmail address of the student to send the email to
    folder_pattern = re.compile(f"^Students-Enrollment-Details.*{re.escape(degree_name)}.*{re.escape(campus)}.*")
    collections = db.list_collection_names()
    collection_name = next((coll for coll in collections if folder_pattern.match(coll)), None)
    
    if collection_name is None:
        return False;
    
    collection = db[collection_name]

    # 查找指定 student_id 的时间表
    student_data = collection.find_one({"StudentID": int(student_id)})
    
    if student_data is None:
        return False
    
    return student_data.get("University Email")